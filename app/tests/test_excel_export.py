from datetime import datetime, timedelta, timezone
from pathlib import Path

import openpyxl

import database
import excel_export


async def test_generate_device_excel(sample_reading, tmp_path, monkeypatch):
    monkeypatch.setattr(excel_export, "EXPORT_DIR", tmp_path)
    path = await excel_export.generate_device_excel("abc123", "TestDevice")
    assert path is not None
    assert path.exists()
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    assert ws.max_row == 2  # header + 1 reading


async def test_generate_device_excel_no_readings(sample_device, tmp_path, monkeypatch):
    monkeypatch.setattr(excel_export, "EXPORT_DIR", tmp_path)
    path = await excel_export.generate_device_excel("abc123", "TestDevice")
    assert path is None


async def test_update_device_excel_creates_if_missing(sample_reading, tmp_path, monkeypatch):
    monkeypatch.setattr(excel_export, "EXPORT_DIR", tmp_path)
    now = datetime.now(timezone.utc)
    path = await excel_export.update_device_excel(
        "abc123", "TestDevice",
        now - timedelta(days=1), now + timedelta(days=1),
    )
    assert path is not None
    wb = openpyxl.load_workbook(path)
    assert wb.active.max_row == 2


async def test_update_device_excel_appends(sample_reading, tmp_path, monkeypatch):
    import aiosqlite

    monkeypatch.setattr(excel_export, "EXPORT_DIR", tmp_path)

    # First: full build (1 reading)
    path = await excel_export.generate_device_excel("abc123", "TestDevice")
    wb = openpyxl.load_workbook(path)
    initial_rows = wb.active.max_row  # header + 1

    # Insert another reading with a distinct timestamp (sample_reading shares the
    # current second; readings are de-duplicated by timestamp now)
    now = datetime.now(timezone.utc)
    new_ts = (now - timedelta(minutes=30)).strftime("%Y-%m-%d %H:%M:%S")
    async with aiosqlite.connect(database.DATABASE_URL) as db:
        await db.execute(
            "INSERT INTO readings (device_id, timestamp, temperature, temp_unit, "
            "gravity, gravity_unit, battery) VALUES (?,?,?,?,?,?,?)",
            ("abc123", new_ts, 22.0, "C", 1.045, "G", 3.8),
        )
        await db.commit()

    path = await excel_export.update_device_excel(
        "abc123", "TestDevice",
        now - timedelta(hours=1), now + timedelta(hours=1),
    )
    wb = openpyxl.load_workbook(path)
    # update appends only the genuinely new reading, not the one already exported
    assert wb.active.max_row == initial_rows + 1


async def test_update_device_excel_no_duplicate_overlap(sample_reading, tmp_path, monkeypatch):
    """A re-run covering an already-exported period must not duplicate rows."""
    monkeypatch.setattr(excel_export, "EXPORT_DIR", tmp_path)

    # Full export, then an update whose window overlaps everything already written.
    await excel_export.generate_device_excel("abc123", "TestDevice")
    now = datetime.now(timezone.utc)
    path = await excel_export.update_device_excel(
        "abc123", "TestDevice",
        now - timedelta(days=1), now + timedelta(hours=1),
    )
    wb = openpyxl.load_workbook(path)
    assert wb.active.max_row == 2  # header + the single original reading, no dupes


async def test_get_device_excel_path_none(tmp_path, monkeypatch):
    monkeypatch.setattr(excel_export, "EXPORT_DIR", tmp_path)
    assert excel_export.get_device_excel_path("NoDevice") is None
