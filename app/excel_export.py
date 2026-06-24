import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime, timezone
from pathlib import Path
import logging

# Timestamps are written to column 1 in this format (UTC+7).
_TS_FMT = '%Y-%m-%d %H:%M:%S'

import database
from config import TZ_UTC7

logger = logging.getLogger(__name__)

EXPORT_DIR = Path("/data/exports")

HEADERS = [
    "Timestamp (UTC+7)",
    "Temperature",
    "Unit",
    "Gravity (SG)",
    "Battery (V)",
    "RSSI",
    "Angle",
    "Interval (s)"
]


def _reading_to_row(reading: dict) -> list:
    """Convert a reading dict to a list of cell values."""
    ts = reading['timestamp']
    if isinstance(ts, str):
        ts = datetime.fromisoformat(ts.replace('Z', '+00:00'))
    if ts.tzinfo:
        ts = ts.astimezone(TZ_UTC7)
    else:
        ts = ts.replace(tzinfo=timezone.utc).astimezone(TZ_UTC7)

    return [
        ts.strftime('%Y-%m-%d %H:%M:%S'),
        reading['temperature'],
        reading.get('temp_unit', 'C'),
        reading['gravity'],
        reading['battery'],
        reading.get('rssi'),
        reading.get('angle'),
        reading.get('interval_sec'),
    ]


def _get_filepath(device_name: str) -> Path:
    """Get Excel file path for a device."""
    safe_name = device_name.replace(' ', '_').replace('/', '_')
    return EXPORT_DIR / f"{safe_name}_data.xlsx"


def _auto_width(ws):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2


async def generate_device_excel(device_id: str, device_name: str) -> Path | None:
    """Generate Excel file with all device data (full rebuild)."""
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    readings = await database.get_all_readings_for_device(device_id)
    if not readings:
        logger.warning(f"No readings found for device {device_name}")
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Device Data"

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row_idx, reading in enumerate(readings, 2):
        for col_idx, value in enumerate(_reading_to_row(reading), 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    _auto_width(ws)

    filepath = _get_filepath(device_name)
    wb.save(filepath)

    logger.info(f"Excel full export saved to {filepath} ({len(readings)} rows)")
    return filepath


def _existing_timestamps(ws) -> set:
    """Collect the timestamp strings (column 1) already present in the sheet."""
    stamps = set()
    for (value,) in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if value is not None:
            stamps.add(str(value))
    return stamps


def _last_timestamp_utc(stamps: set) -> datetime | None:
    """Parse the latest UTC+7 timestamp string back to an aware UTC datetime."""
    if not stamps:
        return None
    try:
        latest = max(stamps)  # ISO-like strings sort chronologically
        return datetime.strptime(latest, _TS_FMT).replace(tzinfo=TZ_UTC7).astimezone(timezone.utc)
    except ValueError:
        return None


async def update_device_excel(device_id: str, device_name: str, start_utc: datetime, end_utc: datetime) -> Path | None:
    """Append new readings up to end_utc to the existing Excel file.

    If the file doesn't exist, creates it with all historical data. Rows are
    de-duplicated by timestamp against what's already in the sheet, so the overlap
    between the initial full export and the first incremental append is not
    double-counted, and a re-run for the same day is a no-op. The lower bound is
    extended back to the last row already in the file, so a missed day (scheduler
    didn't run) gets backfilled on the next run instead of leaving a gap.
    """
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    filepath = _get_filepath(device_name)

    if not filepath.exists():
        logger.info(f"No existing Excel for {device_name}, generating full export")
        return await generate_device_excel(device_id, device_name)

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    existing = _existing_timestamps(ws)
    # Backfill from wherever the file currently ends (covers missed days), but never
    # later than the requested window start.
    last_utc = _last_timestamp_utc(existing)
    effective_start = min(start_utc, last_utc) if last_utc else start_utc

    candidates = await database.get_readings_for_device_period(device_id, effective_start, end_utc)
    new_rows = [row for r in candidates if (row := _reading_to_row(r))[0] not in existing]

    if not new_rows:
        logger.info(f"No new readings for {device_name} in period, Excel unchanged")
        return filepath

    next_row = ws.max_row + 1
    for row in new_rows:
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=next_row, column=col_idx, value=value)
        next_row += 1

    _auto_width(ws)
    wb.save(filepath)

    logger.info(f"Excel updated for {device_name}: appended {len(new_rows)} rows (total {ws.max_row - 1})")
    return filepath


def get_device_excel_path(device_name: str) -> Path | None:
    """Get path to existing Excel file for a device."""
    filepath = _get_filepath(device_name)
    return filepath if filepath.exists() else None
